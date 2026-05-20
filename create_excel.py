#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 팀원 A 데이터
team_a_data = {
    '필드': ['인증/회원', '인증/회원', '인증/회원', '회원관리', '알림', '알림', '알림'],
    '기능명': ['일반 로그인', '회원가입', '소셜 로그인', '프로필 설정', '알림 목록 조회', '알림 읽음 처리', '알림 삭제'],
    '페이지경로': ['/login', '/register', '/login', '/profile/setup', '/notifications', '/notifications', '/notifications'],
    '주요 구현 내용': [
        '이메일, 비밀번호 입력 및 검증, 토큰 발급',
        '이메일, 비밀번호, 닉네임 입력, 약관 동의 처리',
        '구글, 카카오, 깃허브 등 소셜 계정 연동',
        '회원가입 후 프로필 기본정보 설정',
        '좋아요, 댓글, 팔로우, 공지사항 알림 리스트 표시',
        '개별 알림 읽음/읽지 않음 상태 토글',
        '선택 알림 또는 일괄 삭제 기능'
    ],
    '비고': [
        '유효성 검사 필수',
        '입력값 유효성 검사, 중복 확인',
        'OAuth 2.0 기반 구현',
        '닉네임, 자기소개, 프로필 사진',
        '시간순 정렬, 페이지네이션',
        '실시간 반영 필요',
        'soft delete 권장'
    ],
    '우선순위': ['높음', '높음', '중간', '높음', '높음', '중간', '낮음'],
    '상태': ['예정', '예정', '예정', '예정', '예정', '예정', '예정'],
    '협업 사항': [
        '-',
        '팀원 C와 유저 데이터 구조 사전 조율',
        '백엔드 API 연동',
        '팀원 C의 프로필 수정과 일관성 유지',
        '팀원 B, C 이벤트 발생 시 API 연동',
        '푸시 알림 연동 고려',
        '-'
    ]
}

# 팀원 B 데이터
team_b_data = {
    '필드': [
        '피드', '피드', '게시글', '게시글', '게시글', '댓글', '댓글', '댓글', '댓글',
        '채팅', '채팅', '채팅', '채팅', '채팅'
    ],
    '기능명': [
        '피드 목록 조회', '피드 필터링', '게시글 상세 조회', '좋아요 기능', '북마크 기능',
        '댓글 조회', '댓글 작성', '댓글 수정/삭제', '댓글 실시간 반영',
        '채팅방 목록', '채팅방 조회', '메시지 송수신', '채팅방 생성', '메시지 읽음 상태'
    ],
    '페이지경로': [
        '/feed', '/feed', '/post/:id', '/post/:id', '/post/:id',
        '/post/:id', '댓글 모달', '댓글 모달', '댓글 모달',
        '/chat', '/chat/:roomId', '/chat/:roomId', '/chat', '/chat/:roomId'
    ],
    '주요 구현 내용': [
        '무드 기반 게시글 목록 조회, 썸네일 표시',
        '선택된 무드별 게시글 필터링',
        '전체 내용, 좋아요, 댓글, 작성자 정보 표시',
        '게시글 좋아요/취소 및 좋아요 수 업데이트',
        '게시글 북마크/취소, 북마크 목록 저장',
        '댓글 및 대댓글 목록 조회, 중첩 표시',
        '댓글 및 대댓글 작성 기능',
        '본인 댓글만 수정/삭제 가능',
        'WebSocket 기반 실시간 댓글 업데이트',
        '현재 활성 채팅방 목록 표시',
        '채팅방 내 메시지 히스토리 조회',
        'WebSocket + STOMP 기반 실시간 메시지 송수신',
        '1:1 또는 그룹 채팅방 생성',
        '메시지별 읽음/안 읽음 표시'
    ],
    '비고': [
        '무한스크롤 또는 페이지네이션',
        '좋아요 수, 댓글 수, 해시태그 표시',
        '댓글 모달 포함',
        '실시간 반영 필요',
        '마이페이지 저장 목록과 연동',
        '시간순 또는 인기순 정렬',
        '입력값 검증 필수',
        'soft delete 권장',
        '모달 내에서 즉시 반영',
        '최신 메시지 미리보기 포함',
        '페이지네이션으로 이전 메시지 로드',
        '메시지 타입별 처리 (텍스트, 이미지 등)',
        '사용자 선택 및 초대 기능',
        '실시간 동기화 필요'
    ],
    '우선순위': ['높음', '높음', '높음', '높음', '중간', '높음', '높음', '중간', '높음', '높음', '높음', '높음', '높음', '중간'],
    '상태': ['예정']*14,
    '협업 사항': [
        '팀원 C 게시글 데이터 구조 공유 필요',
        '팀원 D 블라인드 상태값(status) 반영',
        '팀원 C와 UI 공동 작업',
        '팀원 A 좋아요 알림 API 연동',
        '팀원 C와 협업',
        '실시간 반영 필요',
        '팀원 A 댓글 알림 API 연동',
        '권한 검증 필요',
        'WebSocket 서버 구성 필요',
        '팀원 C와 UI 협업',
        '팀원 C UI 지원',
        '메시지 암호화 고려',
        '팀원 A 팔로우 정보 활용 가능',
        '-'
    ]
}

# 팀원 C 데이터
team_c_data = {
    '필드': [
        '메인', '무드 선택', '무드 선택', '게시글', '게시글', '게시글', '게시글', '게시글',
        '탐색', '탐색', '탐색', '탐색', '프로필', '프로필', '프로필', '프로필', '프로필', 'UI/UX', 'UI/UX'
    ],
    '기능명': [
        '서비스 소개 화면', '감정별 무드 카드 UI', '무드 선택 값 전달', '게시글 작성 에디터', '이미지 업로드',
        '무드 및 해시태그 설정', '임시저장 기능', '게시글 발행',
        '검색 기능', '해시태그 검색', '인기 태그 조회', '탐색 화면 UI',
        '마이 프로필 조회', '프로필 수정', '내 게시글 조회', '무드 통계 시각화', '팔로우/언팔로우',
        '댓글 UI 지원', '채팅 UI 지원'
    ],
    '페이지경로': [
        '/', '/mood', '/mood → /feed', '/post/create', '/post/create',
        '/post/create', '/post/create', '/post/create',
        '/explore', '/explore', '/explore', '/explore',
        '/profile/:userId', '/profile/:userId/edit', '/profile/:userId', '/profile/:userId', '/profile/:userId',
        '댓글 모달', '/chat'
    ],
    '주요 구현 내용': [
        '서비스 소개, 시작 버튼, 로그인 유도',
        '6가지 무드 카드 제작 및 선택 기능',
        '선택한 무드를 피드 페이지로 전달',
        '제목, 내용, 이미지 업로드 에디터',
        '이미지 업로드 및 미리보기 기능',
        '게시글 무드 선택 및 해시태그 입력',
        '작성 중인 게시글 임시저장',
        '작성 완료된 게시글 발행',
        '키워드 기반 통합 검색 (유저, 해시태그, 키워드)',
        '특정 해시태그 관련 게시글 조회',
        '현재 인기있는 태그 목록 표시',
        '게시글, 해시태그, 유저 탭 구성',
        '프로필 정보, 내 게시글 모아보기',
        '닉네임, 자기소개, 프로필 사진 수정',
        '개인 게시글 목록 및 무드별 필터링',
        '월별, 무드별 통계 차트 표시',
        '유저 팔로우 및 팔로잉 관리',
        '팀원 B의 댓글/대댓글 UI 및 화면 구성 지원',
        '팀원 B의 기본 채팅 화면 구성 및 UI 지원'
    ],
    '비고': [
        '랜딩 페이지 역할',
        '인터랙티브 디자인',
        '상태 관리(Context/Redux)',
        'Rich text editor 사용 권장',
        '파일 크기/형식 검증',
        '자동완성 기능 고려',
        '로컬스토리지 또는 DB 저장',
        '발행 시간 기록',
        '자동완성, 인기검색어',
        '태그별 통계 표시',
        '트렌딩 섹션',
        '반응형 디자인 필수',
        '편집 버튼 포함',
        '저장 전 미리보기 기능',
        '타임라인 방식 표시',
        'Recharts 또는 Chart.js',
        '팔로우 목록 조회',
        '컴포넌트 공유',
        '반응형 디자인 필수'
    ],
    '우선순위': ['높음', '높음', '높음', '높음', '높음', '높음', '중간', '높음', '높음', '중간', '중간', '높음', '높음', '높음', '높음', '중간', '높음', '높음', '높음'],
    '상태': ['예정']*19,
    '협업 사항': [
        '-',
        '팀원 B와 데이터 형식 일치',
        '팀원 B 피드 필터링과 연동',
        '팀원 D 콘텐츠 모니터링과 연동',
        'S3 또는 CDN 연동 필요',
        '팀원 D 해시태그 관리와 연동',
        '-',
        '팀원 B 피드에 즉시 반영',
        '팀원 D와 검색 데이터 구조 조율',
        '팀원 D 통계 시각화와 연동',
        '팀원 D 통계 데이터 활용',
        '팀원 B 피드 컴포넌트 재사용',
        '팀원 A 프로필 설정과 일관성 유지',
        '팀원 A와 데이터 구조 조율',
        '팀원 B 게시글 조회와 일치',
        '팀원 D와 라이브러리 통일',
        '팀원 A 팔로우 알림 API 연동',
        '팀원 B와 협업',
        '팀원 B와 협업'
    ]
}

# 팀원 D 데이터
team_d_data = {
    '필드': [
        '관리자', '회원관리', '회원관리', '회원관리', '콘텐츠관리', '콘텐츠관리', '콘텐츠관리',
        '콘텐츠관리', '콘텐츠관리', '콘텐츠관리', '신고관리', '신고관리', '신고관리',
        '통계', '통계', '통계', '통계', '통계'
    ],
    '기능명': [
        '대시보드', '회원 검색', '회원 상태 관리', '회원 권한 변경', '게시글 모니터링', '댓글 모니터링',
        '이미지 모니터링', '콘텐츠 숨김', '콘텐츠 삭제', '콘텐츠 복구', '신고 내역 검토', '신고 처리',
        '제재 처리', 'DAU 통계', 'MAU 통계', '트래픽 통계', '신고 처리율 통계', '무드 분포 통계'
    ],
    '페이지경로': [
        '/admin/dashboard', '/admin/users', '/admin/users', '/admin/users', '/admin/content', '/admin/content',
        '/admin/content', '/admin/content', '/admin/content', '/admin/content', '/admin/reports', '/admin/reports',
        '/admin/reports', '/admin/statistics', '/admin/statistics', '/admin/statistics', '/admin/statistics/detail', '/admin/statistics'
    ],
    '주요 구현 내용': [
        '주요 지표 (DAU, MAU, 신규가입자) 및 추이 표시',
        '이메일, 닉네임, ID로 회원 검색',
        '일반/정지/관리자 상태 변경',
        '일반 유저를 관리자로 승격/강등',
        '모든 게시글 목록 조회 및 모니터링',
        '모든 댓글 조회 및 부적절한 내용 검토',
        '업로드된 이미지 검수 및 문제 이미지 관리',
        '부적절한 콘텐츠 숨김 처리',
        '규정 위반 콘텐츠 영구 삭제',
        '실수로 삭제된 콘텐츠 복구',
        '사용자 신고 목록 및 상세 내용 검토',
        '신고 승인/기각 처리',
        '경고→일시정지→영구정지 프로세스',
        '일일 활성 사용자 수 시각화',
        '월간 활성 사용자 수 시각화',
        '시간대별, 페이지별 트래픽 분석',
        '신고 접수→처리 시간 및 처리율 시각화',
        '무드별 게시글 수 및 사용자 통계'
    ],
    '비고': [
        '실시간 업데이트',
        '고급 필터링 옵션',
        '즉시 적용 필요',
        '권한 기록 남기기',
        '신고 상태 표시',
        '필터링 기능',
        '썸네일 미리보기',
        '복구 가능성 유지',
        '삭제 이유 기록',
        '휴지통 기능',
        '신고 이유별 분류',
        '처리 결과 통보',
        '제재 기록 유지',
        '추이 분석 차트',
        '성장률 표시',
        '히트맵 또는 라인 차트',
        '개선 추이 표시',
        '파이 차트 또는 바 차트'
    ],
    '우선순위': ['높음', '높음', '높음', '중간', '높음', '중간', '중간', '높음', '높음', '낮음', '높음', '높음', '높음', '높음', '높음', '중간', '중간', '낮음'],
    '상태': ['예정']*18,
    '협업 사항': [
        '-',
        '-',
        '팀원 A와 권한 API 규격 정의',
        '팀원 A와 API 규격 조율',
        '팀원 B, C와 게시글 데이터 구조 공유',
        '팀원 B와 댓글 데이터 구조 공유',
        '팀원 C와 이미지 저장소 정보 공유',
        '팀원 B 피드 조회에 반영 필요',
        '-',
        '-',
        '팀원 A, B, C와 신고 데이터 공유',
        '-',
        '팀원 A와 제재 상태값 규격 조율',
        '차트 라이브러리 통일 필요',
        '팀원 C와 라이브러리 통일',
        '-',
        '-',
        '팀원 C와 협업'
    ]
}

# 엑셀 파일 생성
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# 스타일 정의
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

def create_sheet(wb, sheet_name, data):
    """시트 생성 및 데이터 추가"""
    ws = wb.create_sheet(sheet_name)
    
    # DataFrame 생성
    df = pd.DataFrame(data)
    
    # 데이터 입력
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = border
            cell.alignment = cell_alignment
            
            # 헤더 스타일
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
    
    # 열 너비 설정
    column_widths = [12, 15, 20, 30, 20, 12, 8, 25]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # 행 높이 설정
    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 40

# 각 팀원 시트 생성
create_sheet(wb, '팀원 A - 로그인 회원 알림', team_a_data)
create_sheet(wb, '팀원 B - 피드 댓글 채팅', team_b_data)
create_sheet(wb, '팀원 C - 메인 게시글 탐색', team_c_data)
create_sheet(wb, '팀원 D - 관리자 운영', team_d_data)

# 엑셀 파일 저장
output_path = '/Users/leebc/Library/Mobile Documents/com~apple~CloudDocs/MAC/MoodCast/docs/MoodCast_기능명세서.xlsx'
wb.save(output_path)

print(f"✅ 엑셀 파일 생성 완료!")
print(f"📁 저장 경로: {output_path}")
