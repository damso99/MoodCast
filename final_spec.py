#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

df = pd.read_csv('/Users/leebc/Library/Mobile Documents/com~apple~CloudDocs/MAC/MoodCast/docs/spec_data.tsv', sep='\t')

wb = Workbook()
ws = wb.active
ws.title = 'MoodCast 기능명세서'

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

category_colors = {'공통 UI': 'FFF2CC', '회원': 'FFE6E6', '피드': 'E6F2FF', '게시글': 'E6F2FF', '저장': 'E6F2FF', '검색': 'E6FFE6', '채팅': 'E6F2FF', '프로필': 'E6FFE6', '통계': 'FFF4E6', '설정': 'FFE6E6', '관리자': 'FFF4E6'}

for r_idx, row in enumerate(df.values, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx+1, column=c_idx)
        cell.value = value
        cell.border = border
        cell.alignment = cell_alignment
        
        category_name = row[0]
        if category_name in category_colors:
            cell.fill = PatternFill(start_color=category_colors[category_name], end_color=category_colors[category_name], fill_type='solid')
        if c_idx == 5:
            cell.alignment = center_alignment

for c_idx, col_name in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=c_idx)
    cell.value = col_name
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

column_widths = [10, 15, 12, 35, 8, 10, 15, 10, 20]
for idx, width in enumerate(column_widths, 1):
    col_letter = chr(64 + idx)
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[1].height = 25
for row_idx in range(2, len(df) + 2):
    ws.row_dimensions[row_idx].height = 35

output_path = '/Users/leebc/Library/Mobile Documents/com~apple~CloudDocs/MAC/MoodCast/docs/MoodCast_기능명세서.xlsx'
wb.save(output_path)

print(f"✅ 엑셀 파일 생성 완료!")
print(f"📁 저장 경로: {output_path}")
print(f"📊 총 {len(df)} 개의 기능이 등록되었습니다.")
